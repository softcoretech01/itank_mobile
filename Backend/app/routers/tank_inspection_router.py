# app/routers/tank_inspection_router.py
from fastapi import APIRouter, HTTPException, Depends, status, UploadFile, File, Header
from fastapi.responses import JSONResponse
from fastapi.encoders import jsonable_encoder
from pydantic import BaseModel, Field
from datetime import datetime
from typing import List, Optional, Generator, Any
from sqlalchemy import func, text, inspect
from sqlalchemy.orm import Session
import os
import uuid
import logging
import traceback
import jwt  # PyJWT
import pymysql
from pymysql.cursors import DictCursor
from decimal import Decimal
import urllib.parse
import importlib

from app.utils.s3_utils import AWS_S3_BUCKET
from app.database import get_db, get_db_connection
from app.routers import to_do_list_router
from app.routers.tank_checkpoints_router import FAULTY_STATUS_IDS
from app.models.tank_certificate import TankCertificate
from app.models.inspection_history_model import InspectionHistory

try:
    from PIL import Image
except Exception:
    Image = None

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

router = APIRouter(prefix="/api/tank_inspection_checklist", tags=["tank_inspection"])

UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)
IMAGES_ROOT_DIR = os.path.join(UPLOAD_DIR, "tank_images_mobile")
if not os.path.exists(IMAGES_ROOT_DIR):
    os.makedirs(IMAGES_ROOT_DIR, exist_ok=True)

JWT_SECRET = os.getenv("JWT_SECRET", "change_this_in_production")
JWT_ALGORITHM = os.getenv("JWT_ALGORITHM", "HS256")


def _is_blank_or_zero(v):
    """Return True if value is None, empty string, or numeric 0 (or "0")."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    try:
        return int(v) == 0
    except Exception:
        return False


# Response helpers (uniform envelope)
from fastapi.encoders import jsonable_encoder

def success_resp(message: str, data: Any = None, status_code: int = 200):
    return JSONResponse(
        status_code=status_code,
        content={
            "success": True,
            "message": message,
            "data": jsonable_encoder(data or {})
        },
    )

def error_resp(message: str, status_code: int = 400):
    return JSONResponse(
        status_code=status_code,
        content={"success": False, "message": message, "data": {}},
    )

# -------------------------------------------------------------------
# AUTH – FIXED AND STABLE
# -------------------------------------------------------------------
def get_current_user(
    authorization: Optional[str] = Header(None, alias="Authorization"),
    db: Session = Depends(get_db)
):
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization required")

    token = authorization.replace("Bearer", "").strip()

    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    if not User:
        raise HTTPException(status_code=401, detail="User model not available")

    # 🔥 FIX: Always resolve user from DB
    user = None

    if payload.get("emp_id"):
        user = db.query(User).filter(User.emp_id == int(payload["emp_id"])).first()

    if not user and payload.get("login_name"):
        user = db.query(User).filter(User.login_name == payload["login_name"]).first()

    if not user and payload.get("email"):
        user = db.query(User).filter(User.email == payload["email"]).first()

    if not user and payload.get("sub"):
        user = db.query(User).filter(User.login_name == payload["sub"]).first()

    if not user:
        raise HTTPException(status_code=401, detail="User not found")

    # 🔐 SESSION VALIDATION (CRITICAL)
    session = db.execute(
        text("""
            SELECT 1 FROM login_sessions
            WHERE emp_id = :eid AND still_logged_in = 1
            LIMIT 1
        """),
        {"eid": user.emp_id},
    ).fetchone()

    if not session:
        raise HTTPException(status_code=401, detail="Session expired or logged out")

    return user

# -------------------------------------------------------------------
# EMP_ID RESOLVER (FINAL)
# -------------------------------------------------------------------
def resolve_emp_id(current_user):
    if current_user and hasattr(current_user, "emp_id"):
        return int(current_user.emp_id)
    return None
# -------------------------
# File helpers
# -------------------------
import time
from io import BytesIO
from app.utils.s3_utils import build_s3_key, upload_fileobj_to_s3


def _save_lifter_file(file: UploadFile, tank_number: str, inspection_id: int):
    safe_tank = tank_number or f"inspection_{inspection_id}"
    safe_tank = safe_tank.lower().replace(" ", "_")

    ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
    ts = int(time.time())

    logical_name = f"{safe_tank}_lifter_weight_{ts}{ext}"

    # ---- read file once ----
    buffer = BytesIO()
    while True:
        chunk = file.file.read(64 * 1024)
        if not chunk:
            break
        buffer.write(chunk)
    buffer.seek(0)

    original_key = build_s3_key(logical_name)

    # ✅ USE SAME WORKING HELPER
    upload_fileobj_to_s3(buffer, original_key, file.content_type)

    thumb_key = original_key

    if Image:
        try:
            buffer.seek(0)
            img = Image.open(buffer)
            img.thumbnail((400, 400))
            thumb_buffer = BytesIO()
            img.convert("RGB").save(thumb_buffer, format="JPEG")
            thumb_buffer.seek(0)

            thumb_name = f"{safe_tank}_lifter_weight_{ts}_thumb.jpg"
            thumb_key = build_s3_key(thumb_name)

            upload_fileobj_to_s3(thumb_buffer, thumb_key, "image/jpeg")
        except Exception as e:
            logger.warning("Thumbnail generation failed", exc_info=True)

    return {
        "image_path": original_key,
        "thumbnail_path": thumb_key
    }



def fetch_pi_next_inspection_date(db: Session, tank_id: int):
    try:
        row = db.execute(
            text(
                """
                SELECT next_insp_date
                FROM tank_certificate
                WHERE tank_id = :tank_id
                ORDER BY next_insp_date IS NULL ASC, next_insp_date DESC
                LIMIT 1
                """
            ),
            {"tank_id": tank_id},
        ).fetchone()
        print(f"Debug: tank_id={tank_id}, row={row}")
        if not row:
            print("Debug: no row")
            return None
        try:
            if hasattr(row, "_mapping"):
                mapping = row._mapping
                result = next(iter(mapping.values()), None)
            elif isinstance(row, dict):
                result = next(iter(row.values()), None)
            else:
                result = row[0]
            print(f"Debug: result={result}")
            return result
        except Exception as e:
            print(f"Debug: exception in parsing row: {e}")
            try:
                result = list(row.values())[0]
                print(f"Debug: fallback result={result}")
                return result
            except Exception as e2:
                print(f"Debug: fallback exception: {e2}")
                return None
    except Exception as exc:
        print(f"Debug: exception in fetch: {exc}")
        logger.warning("Could not fetch PI next inspection date for tank_id %s: %s", tank_id, exc)
        return None


def generate_report_number(db: Session, inspection_date: datetime, inspection_type_id: Optional[int] = None) -> str:
    date_str = inspection_date.strftime("%d%m%Y")
    
    # Determine Prefix
    prefix_part = "SG-T1" # Default
    if inspection_type_id:
        try:
            it_row = db.execute(text("SELECT inspection_type_name FROM inspection_type WHERE id = :id"), {"id": inspection_type_id}).fetchone()
            if it_row:
                # Handle mapping or tuple access
                if hasattr(it_row, "_mapping"):
                    itype_name = it_row._mapping.get("inspection_type_name", "").upper()
                else:
                    itype_name = it_row[0].upper()
                
                # Normalize: remove hyphens and spaces to match user request (e.g. "On-Hire" -> "ONHIRE")
                normalized_name = itype_name.replace("-", "").replace(" ", "")

                # Check for special types
                if normalized_name in ["ONHIRE", "OFFHIRE", "CONDITION"]:
                    prefix_part = f"SG-{normalized_name}-T1"
        except Exception as e:
            logger.warning(f"Error fetching inspection type name for report number generation: {e}")

    for attempt in range(3):
        try:
            cnt_row = db.execute(
                text("SELECT COUNT(*) AS cnt FROM tank_inspection_details WHERE DATE(inspection_date) = :d"),
                {"d": inspection_date.date()},
            ).fetchone()
            if cnt_row is None:
                count = 0
            else:
                if hasattr(cnt_row, "_mapping"):
                    count = int(cnt_row._mapping.get("cnt", 0))
                elif isinstance(cnt_row, dict):
                    count = int(cnt_row.get("cnt", 0))
                else:
                    count = int(cnt_row[0])
        except Exception:
            count = 0

        next_counter = (count or 0) + 1
        report_number = f"{prefix_part}-{date_str}-{next_counter:02d}"

        try:
            existing = db.execute(text("SELECT 1 FROM tank_inspection_details WHERE report_number = :rn LIMIT 1"), {"rn": report_number}).fetchone()
            if not existing:
                return report_number
        except Exception:
            return report_number

        logger.warning(f"Report number collision for {report_number}, retrying...")

    raise RuntimeError(f"Unable to generate unique report number after retries for date {date_str}")


def fetch_tank_details(db: Session, tank_number: str):
    result = db.execute(
        text(
            """
            SELECT working_pressure, frame_type, design_temperature, cabinet_type, mfgr, ownership
            FROM tank_details
            WHERE tank_number = :tank_number
            LIMIT 1
            """
        ),
        {"tank_number": tank_number},
    ).fetchone()

    if not result:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tank details not found for tank_number: {tank_number}",
        )

    try:
        if hasattr(result, "_mapping"):
            mapping = result._mapping
            working_pressure = mapping.get("working_pressure", None)
            frame_type = mapping.get("frame_type", None)
            design_temperature = mapping.get("design_temperature", None)
            cabinet_type = mapping.get("cabinet_type", None)
            mfgr = mapping.get("mfgr", None)
            ownership = mapping.get("ownership", None)
        else:
            working_pressure = result[0]
            frame_type = result[1]
            design_temperature = result[2]
            cabinet_type = result[3]
            mfgr = result[4]
            ownership = result[5]
    except Exception:
        try:
            rowm = dict(result)
            working_pressure = rowm.get("working_pressure")
            frame_type = rowm.get("frame_type")
            design_temperature = rowm.get("design_temperature")
            cabinet_type = rowm.get("cabinet_type")
            mfgr = rowm.get("mfgr")
            ownership = rowm.get("ownership")
        except Exception:
            working_pressure = frame_type = design_temperature = cabinet_type = mfgr = ownership = None

    def to_float_if_decimal(val):
        if isinstance(val, Decimal):
            return float(val)
        return val

    return {
        "working_pressure": working_pressure,
        "cabinet_type": cabinet_type,
        "frame_type": frame_type,
        "design_temperature": design_temperature,
        "mfgr": mfgr,
        "ownership": ownership,
    }


# -------------------------
# Pydantic schemas (updated to use tank_id in create/update)
# -------------------------
class TankInspectionCreate(BaseModel):
    tank_id: int = Field(..., description="tank_details.tank_id (client must send tank_id)")
    status_id: Optional[int] = None
    product_id: Optional[int] = None
    inspection_type_id: Optional[int] = None
    location_id: Optional[int] = None
    safety_valve_brand_id: Optional[int] = None
    safety_valve_model_id: Optional[int] = None  # nullable
    safety_valve_size_id: Optional[int] = None   # nullable
    notes: Optional[str] = None
    operator_id: Optional[int] = None   # manual operator id entered by user (nullable)
    vacuum_reading: Optional[str] = None
    lifter_weight_value: Optional[str] = None

    class Config:
        json_schema_extra = {
            "example": {
                "tank_id": 0,
                "status_id": 0,
                "product_id": 0,
                "inspection_type_id": 0,
                "location_id": 0,
                "safety_valve_brand_id": 0,
                "safety_valve_model_id": 0,
                "safety_valve_size_id": 0,
                "notes": "All checks ok",
                "operator_id": 0
            }
        }

class TankInspectionResponse(BaseModel):
    inspection_id: int
    tank_number: str
    report_number: str
    inspection_date: datetime
    status_id: Optional[int] = None
    product_id: Optional[int] = None
    inspection_type_id: Optional[int] = None
    location_id: Optional[int] = None
    working_pressure: Optional[float] = None
    design_temperature: Optional[str] = None
    frame_type: Optional[str] = None
    cabinet_type: Optional[str] = None
    mfgr: Optional[str] = None
    pi_next_inspection_date: Optional[str] = None
    safety_valve_brand_id: Optional[int] = None
    safety_valve_model_id: Optional[int] = None
    safety_valve_size_id: Optional[int] = None
    notes: Optional[str] = None
    created_by: Optional[str] = None
    operator_id: Optional[int] = None
    emp_id: int     # NOT optional - must be the logged-in user's ID
    ownership: Optional[str] = None
    lifter_weight: Optional[str] = None
    vacuum_reading: Optional[str] = None
    lifter_weight_value: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class TankInspectionUpdate(BaseModel):
    inspection_date: Optional[datetime] = None
    # client can send tank_id if they want to change which tank this inspection refers to (rare) --
    # if provided, code resolves tank_number from tank_id.
    tank_id: Optional[int] = None
    status_id: Optional[int] = None
    inspection_type_id: Optional[int] = None
    product_id: Optional[int] = None
    location_id: Optional[int] = None
    safety_valve_brand_id: Optional[int] = None
    safety_valve_model_id: Optional[int] = None      # nullable
    safety_valve_size_id: Optional[int] = None       # nullable

    class Config:
        from_attributes = True


# -------------------------
# NEW: GET ALL INSPECTIONS
# -------------------------
@router.get("/list/all")
def get_all_inspections(
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Fetch all inspection records from tank_inspection_details.
    Includes joins for human-readable names.
    """
    try:
        # Resolve emp_id for filtering if needed, but usually list all for admin/management
        # For now, let's list all as requested: "all the created"
        
        query = text("""
            SELECT 
                ti.*,
                ps.status_name,
                pm.product_name,
                it.inspection_type_name,
                l.location_name,
                sv.brand_name AS safety_valve_brand_name
            FROM tank_inspection_details ti
            LEFT JOIN tank_status ps ON ti.status_id = ps.id
            LEFT JOIN product_master pm ON ti.product_id = pm.id
            LEFT JOIN inspection_type it ON ti.inspection_type_id = it.id
            LEFT JOIN location_master l ON ti.location_id = l.id
            LEFT JOIN safety_valve_brand sv ON ti.safety_valve_brand_id = sv.id
            ORDER BY ti.inspection_date DESC, ti.inspection_id DESC
        """)
        
        results = db.execute(query).mappings().fetchall()
        
        # Convert Decimals and datetimes for JSON, and remove IDs
        def clean_row(r):
            d = dict(r)
            # Remove IDs as requested to return names "instead of" ids
            for id_key in ["status_id", "product_id", "inspection_type_id", "location_id", "safety_valve_brand_id"]:
                d.pop(id_key, None)
                
            for k, v in d.items():
                if isinstance(v, Decimal):
                    d[k] = float(v)
                elif isinstance(v, datetime):
                    d[k] = v.isoformat()
            return d

        data = [clean_row(r) for r in results]
        return success_resp("All inspections fetched", data)

    except Exception as e:
        logger.error(f"Error fetching all inspections: {e}", exc_info=True)
        return error_resp(f"Internal server error: {str(e)}", 500)


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from fastapi.responses import StreamingResponse

# -------------------------
# EXPORT INSPECTIONS TO EXCEL
# -------------------------
@router.get("/export-to-excel")
def export_inspections_to_excel(
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Export all inspection records to an Excel file with detailed information.
    """
    try:
        query = text("""
            SELECT 
                ti.inspection_id,
                ti.report_number,
                ti.tank_number,
                ti.inspection_date,
                ti.mfgr,
                ti.working_pressure,
                ti.design_temperature,
                ti.frame_type,
                ti.cabinet_type,
                ti.pi_next_inspection_date,
                ti.notes,
                ti.vacuum_reading,
                ti.lifter_weight_value,
                ti.ownership,
                ti.operator_id,
                ti.is_submitted,
                ti.web_submitted,
                ti.created_by,
                ti.created_at,
                ti.updated_by,
                ti.updated_at,
                ps.status_name,
                it.inspection_type_name,
                pl.location_name,
                pm.product_name,
                sb.brand_name AS safety_valve_brand_name,
                sm.model_name AS safety_valve_model_name,
                ss.size_label AS safety_valve_size_name
            FROM tank_inspection_details ti
            LEFT JOIN tank_status ps ON ti.status_id = ps.id
            LEFT JOIN inspection_type it ON ti.inspection_type_id = it.id
            LEFT JOIN location_master pl ON ti.location_id = pl.id
            LEFT JOIN product_master pm ON ti.product_id = pm.id
            LEFT JOIN safety_valve_brand sb ON ti.safety_valve_brand_id = sb.id
            LEFT JOIN safety_valve_model sm ON ti.safety_valve_model_id = sm.id
            LEFT JOIN safety_valve_size ss ON ti.safety_valve_size_id = ss.id
            ORDER BY ti.inspection_date DESC
        """)
        
        results = db.execute(query).mappings().fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspection Reports"
        
        headers = [
            "ID", "Report No", "Tank No", "Date", "Status", "Inspection Type", "Location", "Product", 
            "Manufacturer", "Working Pressure", "Design Temp", "Frame Type", "Cabinet Type", "Next Inspection", 
            "SV Brand", "SV Model", "SV Size", "Vacuum Reading", "Lifter Weight", "Ownership", "Operator ID", 
            "Is Submitted", "Web Submitted", "Created By", "Created At", "Updated By", "Updated At"
        ]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
        for row_num, row_data in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=row_data.get("inspection_id"))
            ws.cell(row=row_num, column=2, value=row_data.get("report_number"))
            ws.cell(row=row_num, column=3, value=row_data.get("tank_number"))
            ws.cell(row=row_num, column=4, value=str(row_data.get("inspection_date")))
            ws.cell(row=row_num, column=5, value=row_data.get("status_name"))
            ws.cell(row=row_num, column=6, value=row_data.get("inspection_type_name"))
            ws.cell(row=row_num, column=7, value=row_data.get("location_name") or "-")
            ws.cell(row=row_num, column=8, value=row_data.get("product_name") or "-")
            ws.cell(row=row_num, column=9, value=row_data.get("mfgr") or "-")
            ws.cell(row=row_num, column=10, value=row_data.get("working_pressure") or "-")
            ws.cell(row=row_num, column=11, value=row_data.get("design_temperature") or "-")
            ws.cell(row=row_num, column=12, value=row_data.get("frame_type") or "-")
            ws.cell(row=row_num, column=13, value=row_data.get("cabinet_type") or "-")
            ws.cell(row=row_num, column=14, value=row_data.get("pi_next_inspection_date") or "-")
            ws.cell(row=row_num, column=15, value=row_data.get("safety_valve_brand_name") or "-")
            ws.cell(row=row_num, column=16, value=row_data.get("safety_valve_model_name") or "-")
            ws.cell(row=row_num, column=17, value=row_data.get("safety_valve_size_name") or "-")
            ws.cell(row=row_num, column=18, value=row_data.get("vacuum_reading") or "-")
            ws.cell(row=row_num, column=19, value=row_data.get("lifter_weight_value") or "-")
            ws.cell(row=row_num, column=20, value=row_data.get("ownership") or "-")
            ws.cell(row=row_num, column=21, value=row_data.get("operator_id") or "-")
            
            # Is Submitted status
            is_sub = "SUBMITTED" if row_data.get("is_submitted") == 1 else "DRAFT"
            ws.cell(row=row_num, column=22, value=is_sub)
            
            # Web Submitted status with UPDATED logic
            web_status = ""
            if row_data.get("web_submitted") == 1:
                web_status = "UPDATED" if row_data.get("is_submitted") == 1 else "SUBMITTED"
            ws.cell(row=row_num, column=23, value=web_status)
            
            ws.cell(row=row_num, column=24, value=row_data.get("created_by"))
            ws.cell(row=row_num, column=25, value=str(row_data.get("created_at")))
            ws.cell(row=row_num, column=26, value=row_data.get("updated_by"))
            ws.cell(row=row_num, column=27, value=str(row_data.get("updated_at")))

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            header_val = str(col[0].value)
            max_length = len(header_val)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Inspection_Reports_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Error exporting inspections: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"success": False, "message": "Failed to export inspections"})


# -------------------------
# Auth helper (kept as before)
# -------------------------
try:
    from app.models.users_model import User
except Exception:
    User = None


def get_current_user(authorization: Optional[str] = Header(None, alias="Authorization"), db: Session = Depends(get_db)):
    if not authorization:
        return None
    auth = authorization.strip()
    token = auth
    if len(auth) >= 6 and auth[:6].lower() == "bearer":
        token_part = auth[6:]
        token = token_part.lstrip(" :\t")
    token = token.strip()
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    except Exception:
        raise HTTPException(status_code=401, detail="Could not validate credentials")

    if User is None:
        return payload

    user = None
    try:
        if "emp_id" in payload and payload["emp_id"] is not None:
            try:
                user = db.query(User).filter(User.emp_id == int(payload["emp_id"])).first()
            except Exception:
                user = db.query(User).filter(User.emp_id == payload["emp_id"]).first()
        elif "email" in payload and payload["email"]:
            user = db.query(User).filter(User.email == payload["email"]).first()
        elif "sub" in payload and payload["sub"]:
            sub = payload["sub"]
            try:
                user = db.query(User).filter((User.email == sub) | (User.emp_id == int(sub))).first()
            except Exception:
                user = db.query(User).filter((User.email == sub) | (User.emp_id == sub)).first()
    except Exception:
        raise HTTPException(status_code=401, detail="Validation error")

    if not user:
        raise HTTPException(status_code=401, detail="User not found")
        # Enforce session validity (check if logged out)
        
    try:
        session_row = db.execute(
            text("SELECT 1 FROM login_sessions WHERE emp_id = :eid AND still_logged_in = 1 LIMIT 1"),
            {"eid": user.emp_id}
        ).fetchone()
        
        if not session_row:
             # User logged out explicitly
             raise HTTPException(status_code=401, detail="Session invalid or logged out")
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"Session check error: {e}")
        # Fail safe to unauthorized if session state cannot be verified
        raise HTTPException(status_code=401, detail="Could not verify session active state")


    return user


@router.get("/auth/debug-token")
def debug_token(authorization: Optional[str] = Header(None, alias="Authorization")):
    if not authorization:
        return error_resp("No Authorization header", 400)
    token = authorization.strip()
    if token.lower().startswith("bearer "):
        token = token.split(" ", 1)[1].strip()
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM], options={"verify_signature": False})
        return success_resp("Decoded token payload (no signature verification)", payload, 200)
    except Exception as e:
        return error_resp(f"Failed to decode token: {e}", 400)


# -------------------------
# Helper: validate operator exists in operators table
# -------------------------
def operator_exists(db: Session, operator_id: int) -> bool:
    try:
        r = db.execute(text("SELECT 1 FROM operators WHERE operator_id = :op LIMIT 1"), {"op": operator_id}).fetchone()
        return bool(r)
    except Exception:
        return False


# -------------------------
# Masters endpoint (kept)
# -------------------------
@router.get("/masters")
def get_all_tank_inspection_masters():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(DictCursor) as cursor:
            masters = {
                "tank_statuses": ("tank_status", ["status_id", "status_name", "description", "created_at", "updated_at"]),
                "products": ("product_master", ["product_id", "product_name", "description", "created_at", "updated_at"]),
                "inspection_types": ("inspection_type", ["inspection_type_id", "inspection_type_name", "description", "created_at", "updated_at"]),
                "locations": ("location_master", ["location_id", "location_name", "description", "created_at", "updated_at"]),
                "safety_valve_brands": ("safety_valve_brand", ["id", "brand_name", "description", "created_at", "updated_at"]),
                "safety_valve_models": ("safety_valve_model", ["id", "model_name", "description", "created_at", "updated_at"]),
                "safety_valve_sizes": ("safety_valve_size", ["id", "size_label", "description", "created_at", "updated_at"]),
            }

            out_data = {}

            for key, (table, expected_fields) in masters.items():
                try:
                    cursor.execute(f"SELECT * FROM `{table}` LIMIT 100")
                    sample_rows = cursor.fetchall() or []
                except Exception as ex:
                    logger.warning("Failed to fetch table %s: %s", table, ex, exc_info=True)
                    out_data[key] = []
                    continue

                real_cols = list(sample_rows[0].keys()) if sample_rows else []
                if not real_cols:
                    try:
                        cursor.execute(f"SELECT * FROM `{table}` LIMIT 0")
                        real_cols = [d[0] for d in cursor.description] if cursor.description else []
                    except Exception:
                        real_cols = []

                def pick_real_col_for_expected(ef):
                    if ef.endswith("_id"):
                        if ef in real_cols:
                            return ef
                        if "id" in real_cols:
                            return "id"
                        base = ef[:-3]
                        if f"{base}_id" in real_cols:
                            return f"{base}_id"
                        if f"{base}id" in real_cols:
                            return f"{base}id"
                        return None
                    candidates = [ef]
                    if ef.endswith("_name"):
                        candidates.append(ef.replace("_name", "name"))
                        candidates.append(ef.replace("_name", ""))
                    for c in candidates:
                        if c in real_cols:
                            return c
                    for c in real_cols:
                        if c.lower().endswith(ef.split("_")[-1].lower()):
                            return c
                    return None

                chosen_map = {ef: pick_real_col_for_expected(ef) for ef in expected_fields}
                mapped = []
                for r in sample_rows:
                    out_row = {}
                    for ef in expected_fields:
                        real = chosen_map.get(ef)
                        val = None
                        if real and real in r:
                            val = r.get(real)
                        out_row[ef] = val
                    mapped.append(out_row)
                out_data[key] = jsonable_encoder(mapped)

            return success_resp("Master data fetched successfully", out_data, 200)
    except Exception as e:
        logger.error(f"Error fetching masters: {e}", exc_info=True)
        return error_resp("Error fetching master data", 500)
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass


# Simple validator for tank existence (kept)
def validate_tank_exists(db: Session, tank_number: str):
    result = db.execute(
        text("SELECT 1 FROM tank_header WHERE tank_number = :tank_number"),
        {"tank_number": tank_number},
    ).fetchone()
    if not result:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Tank not existing: {tank_number}")


# -------------------------
# Active tanks endpoint
# -------------------------
@router.get("/active-tanks")
def get_active_tanks(db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    try:
        rows = db.execute(text("SELECT tank_id, tank_number FROM tank_details WHERE status = 'active'")).mappings().all()
        data = [dict(r) for r in rows]
        return success_resp("Active tanks fetched", {"active_tanks": jsonable_encoder(data)}, 200)
    except Exception as e:
        logger.error(f"Error fetching active tanks: {e}", exc_info=True)
        return error_resp("Error fetching active tanks", 500)


# -------------------------
# Lifter weight thumbnail endpoint
# -------------------------
@router.get("/lifter_weight/{inspection_id}")
def get_lifter_weight_thumbnail(
    inspection_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        # Try to select thumbnail column if it exists
        try:
            row = db.execute(
                text("""
                    SELECT lifter_weight,
                           lifter_weight_thumbnail,
                           tank_number
                    FROM tank_inspection_details
                    WHERE inspection_id = :id
                """),
                {"id": inspection_id}
            ).fetchone()
        except Exception:
            # Fallback for older schema (no thumbnail column)
            row = db.execute(
                text("""
                    SELECT lifter_weight,
                           tank_number
                    FROM tank_inspection_details
                    WHERE inspection_id = :id
                """),
                {"id": inspection_id}
            ).fetchone()

        if not row:
            return error_resp("No inspection found.", 404)

        # Normalize row
        if hasattr(row, "_mapping"):
            m = row._mapping
            tank_number = m.get("tank_number")
            rel_path = m.get("lifter_weight")
            thumb_db = m.get("lifter_weight_thumbnail") if "lifter_weight_thumbnail" in m else None
        else:
            # Index-based fallback
            rel_path = row[0]
            if len(row) == 3:
                thumb_db = row[1]
                tank_number = row[2]
            else:
                thumb_db = None
                tank_number = row[1]

        if not rel_path:
            return error_resp("No lifter weight photo found for this inspection.", 404)

        # If DB has a thumbnail path, return it directly
        if thumb_db:
            return success_resp(
                "Lifter weight thumbnail fetched",
                {
                    "inspection_id": inspection_id,
                    "thumbnail_path": thumb_db
                },
                200,
            )

        # ---------- NEW: derive thumbnail from lifter_weight ----------
        # rel_path example: "uploads/iso_tank/2025/12/1733647560_TANK123_lifter_weight.jpg"
        folder = os.path.dirname(rel_path) or ""
        base_name = os.path.basename(rel_path)
        name_no_ext, _ext = os.path.splitext(base_name)

        # Our _save_lifter_file saves thumbnail as "{timestamp}_{tank}_lifter_weight_thumb.jpg"
        thumb_name = f"{name_no_ext}_thumb.jpg"

        if folder:
            thumb_rel = f"{folder}/{thumb_name}"
        else:
            thumb_rel = thumb_name

        # Check if thumbnail file exists, else fallback to original image
        thumb_abs = os.path.join(os.getcwd(), thumb_rel)
        if not os.path.exists(thumb_abs):
            thumb_rel = rel_path

        return success_resp(
            "Lifter weight thumbnail fetched",
            {
                "inspection_id": inspection_id,
                "thumbnail_path": thumb_rel,
            },
            200,
        )

    except Exception as e:
        logger.error(f"Error fetching lifter weight thumbnail for {inspection_id}: {e}", exc_info=True)
        return error_resp("Internal server error", 500)

# -------------------------
# Create Tank Inspection (flat payload with master ids)
# -------------------------
@router.post("/create/tank_inspection", status_code=status.HTTP_201_CREATED)
def create_tank_inspection(
    payload: TankInspectionCreate,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user),
):
    try:
        # --- Resolve tank_number from payload.tank_id ---
        try:
            tn_row = db.execute(
                text("SELECT tank_number FROM tank_details WHERE tank_id = :tid LIMIT 1"),
                {"tid": payload.tank_id},
            ).fetchone()
        except Exception as e:
            logger.error("DB error resolving tank_number: %s", e, exc_info=True)
            return error_resp(f"Tank not found for id: {payload.tank_id}", 404)

        if not tn_row:
            return error_resp(f"Tank not found for id: {payload.tank_id}", 404)

        # Handle row mapping safely
        if hasattr(tn_row, "_mapping"):
            tank_number = tn_row._mapping.get("tank_number")
        elif isinstance(tn_row, dict):
            tank_number = tn_row.get("tank_number")
        else:
            tank_number = tn_row[0]

        # --- Helper: Strictly check if value is a valid ID ---
        def is_valid_id(val):
            if val is None:
                return False
            if isinstance(val, int) and val > 0:
                return True
            if isinstance(val, str) and val.isdigit() and int(val) > 0:
                return True
            return False

        # --- Validate master ids (Only if provided and non-zero) ---
        # Note: During CREATE, all IDs are allowed to be 0 or null
        # Validation happens at SUBMIT time via validation/submit endpoints
        master_checks = [
            ("tank_status", payload.status_id, "status_id"),
            ("product_master", payload.product_id, "product_id"),
            ("inspection_type", payload.inspection_type_id, "inspection_type_id"),
            ("location_master", payload.location_id, "location_id"),
            # Safety valves (Optional)
            ("safety_valve_brand", payload.safety_valve_brand_id, "safety_valve_brand_id"),
            ("safety_valve_model", payload.safety_valve_model_id, "safety_valve_model_id"),
            ("safety_valve_size", payload.safety_valve_size_id, "safety_valve_size_id"),
        ]

        # --- Prepare Inspection Data ---
        tank_details = fetch_tank_details(db, tank_number)
        inspection_date = datetime.now()
        
        # Resolve Emp ID (from JWT / session when available; never trust payload)
        emp_id_val = resolve_emp_id(current_user)

        if not emp_id_val:
            return error_resp("Unable to resolve emp_id from token", 401)


        # Check for any existing non-submitted inspection for this tank
        active_inspection = db.execute(
            text("""
                SELECT inspection_id, report_number
                FROM tank_inspection_details 
                WHERE tank_number = :tn 
                  AND (is_submitted IS NULL OR is_submitted = 0)
                  AND (web_submitted IS NULL OR web_submitted = 0)
                  AND (is_reviewed IS NULL OR is_reviewed = 0)
                LIMIT 1
            """),
            {"tn": tank_number}
        ).fetchone()

        if active_inspection:
            report_num = ""
            if hasattr(active_inspection, "_mapping"):
                report_num = active_inspection._mapping.get("report_number")
            elif hasattr(active_inspection, "report_number"):
                report_num = active_inspection.report_number
            else:
                try:
                    report_num = active_inspection[1]
                except Exception:
                    report_num = "Unknown"
                    
            if not report_num:
                report_num = "Unknown"
                
            return error_resp(f"An active, non-submitted inspection (Report No: {report_num}) already exists for tank {tank_number}. Please complete or edit the existing inspection before creating a new one.", 400)

        # Duplicate Check
        existing = db.execute(
            text("SELECT inspection_id FROM tank_inspection_details WHERE tank_number = :tn AND DATE(inspection_date) = :d AND inspection_type_id = :itype LIMIT 1"),
            {"tn": tank_number, "d": inspection_date.date(), "itype": payload.inspection_type_id},
        ).fetchone()
        if existing:
            return error_resp("Inspection already exists", 400)

        # Generate Reports
        report_number = generate_report_number(db, inspection_date, inspection_type_id=payload.inspection_type_id)
        pi_next_date = fetch_pi_next_inspection_date(db, payload.tank_id)
        ownership_val = tank_details.get("ownership")
        # Sanitize Safety Valve IDs for Insert (Ensure None if invalid)
        svb = payload.safety_valve_brand_id if is_valid_id(payload.safety_valve_brand_id) else None
        svm = payload.safety_valve_model_id if is_valid_id(payload.safety_valve_model_id) else None
        svs = payload.safety_valve_size_id if is_valid_id(payload.safety_valve_size_id) else None

        # --- INSERT ---
        try:
            db.execute(
                text("""
                    INSERT INTO tank_inspection_details
                    (inspection_date, report_number, tank_number, tank_id, status_id, product_id, inspection_type_id, location_id,
                     working_pressure, frame_type, design_temperature, cabinet_type, mfgr, pi_next_inspection_date,
                     safety_valve_brand_id, safety_valve_model_id, safety_valve_size_id, notes,
                     vacuum_reading, lifter_weight_value,
                     created_by, updated_by,
                     operator_id, emp_id, ownership, is_submitted, created_at, updated_at)
                    VALUES
                    (:inspection_date, :report_number, :tank_number, :tank_id, :status_id, :product_id, :inspection_type_id, :location_id,
                     :working_pressure, :frame_type, :design_temperature, :cabinet_type, :mfgr, :pi_next_inspection_date,
                     :svb, :svm, :svs, :notes,
                     :vacuum_reading, :lifter_weight_value,
                     :created_by, :updated_by,
                     :operator_id, :emp_id, :ownership, :is_submitted, NOW(), NOW())
                """),
                {
                    "inspection_date": inspection_date,
                    "report_number": report_number,
                    "tank_number": tank_number,
                    "tank_id": payload.tank_id,
                    "status_id": None if payload.status_id in [None, 0, "0", ""] else payload.status_id,
                    "product_id": None if payload.product_id in [None, 0, "0", ""] else payload.product_id,
                    "inspection_type_id": payload.inspection_type_id if is_valid_id(payload.inspection_type_id) else None,
                    "location_id": payload.location_id if is_valid_id(payload.location_id) else None,
                    "working_pressure": tank_details.get("working_pressure"),
                    "frame_type": tank_details.get("frame_type"),
                    "design_temperature": tank_details.get("design_temperature"),
                    "cabinet_type": tank_details.get("cabinet_type"),
                    "mfgr": tank_details.get("mfgr"),
                    "pi_next_inspection_date": pi_next_date,
                    "svb": svb, "svm": svm, "svs": svs,
                    "notes": payload.notes,
                    "vacuum_reading": payload.vacuum_reading,
                    "lifter_weight_value": payload.lifter_weight_value,
                    # created_by / updated_by MUST come from logged-in user
                    "created_by": emp_id_val,
                    "updated_by": emp_id_val,
                    "operator_id": payload.operator_id,
                    "emp_id": emp_id_val,
                    "ownership": ownership_val,
                    "is_submitted": 0,
                },
            )
            db.commit()
        except Exception as e:
            db.rollback()
            logger.error("Failed to create tank inspection record: %s", e, exc_info=True)
            return error_resp(f"Internal server error: {e}", 500)

        # --- Return Created Record ---
        new_row = db.execute(text("SELECT * FROM tank_inspection_details WHERE report_number = :rn"), {"rn": report_number}).fetchone()
        
        # Convert row to dict safely
        if hasattr(new_row, "_mapping"):
            out = dict(new_row._mapping)
        elif isinstance(new_row, dict):
            out = new_row
        else:
            out = dict(zip(new_row.keys(), new_row))

        return success_resp("Inspection created successfully", out, 201)

    except Exception as e:
        logger.error(f"Error creating tank inspection: {e}", exc_info=True)
        return error_resp(f"Internal server error: {e}", 500)


# -------------------------
# Update tank_inspection_details (PUT)
# -------------------------
class TankInspectionUpdateModel(BaseModel):
    inspection_date: Optional[datetime] = None
    tank_id: Optional[int] = None
    status_id: Optional[int] = None
    inspection_type_id: Optional[int] = None
    product_id: Optional[int] = None
    location_id: Optional[int] = None
    safety_valve_brand_id: Optional[int] = None
    safety_valve_model_id: Optional[int] = None      # nullable
    safety_valve_size_id: Optional[int] = None       # nullable
    vacuum_reading: Optional[str] = None
    lifter_weight_value: Optional[str] = None

    class Config:
        from_attributes = True


@router.put("/update/tank_inspection_details/{inspection_id}")
def update_tank_inspection_details(
    inspection_id: int, 
    payload: TankInspectionUpdateModel, 
    db: Session = Depends(get_db), 
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        # 1. Check if inspection exists and role permissions
        row = db.execute(text("SELECT is_submitted, web_submitted FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp("Inspection not found", 404)

        is_submitted = int(row[0])
        web_submitted = int(row[1])
        role_id = current_user.role_id
        if (is_submitted == 1 or web_submitted == 1) and role_id == 2:
            return error_resp("Cannot edit submitted inspection", 403)

        params = {"id": inspection_id}
        updates = []

        # Helper: Strictly check if value is a valid ID (int > 0)
        def is_valid_id(val):
            if val is None: return False
            if isinstance(val, int) and val > 0: return True
            if isinstance(val, str) and val.isdigit() and int(val) > 0: return True
            return False

        # Helper: Get set fields safely (Pydantic v1/v2 compat)
        try:
            # Try Pydantic v2
            update_data = payload.model_dump(exclude_unset=True)
        except AttributeError:
            # Fallback to Pydantic v1
            update_data = payload.dict(exclude_unset=True)

        # --- Handle Special Fields (operator_id, emp_id, tank_id) ---
        
        # Operator ID (Optional)
        if "operator_id" in update_data:
            op_id = update_data["operator_id"]
            # Treat 0 as None if needed, or just pass it. Assuming 0 means "no operator" -> None
            if op_id == 0:
                op_id = None
            updates.append("operator_id = :operator_id")
            params["operator_id"] = op_id

        # Emp ID (Auto-resolve)
        emp_id_val = resolve_emp_id(current_user)

        if not emp_id_val:
            return error_resp("Unable to resolve emp_id from token", 401)

        # Tank ID (Resolve Number)
        if "tank_id" in update_data and update_data["tank_id"] is not None:
            tid = update_data["tank_id"]
            tn_row = db.execute(text("SELECT tank_number FROM tank_details WHERE tank_id = :tid LIMIT 1"), {"tid": tid}).fetchone()
            if not tn_row:
                return error_resp(f"Tank not found for id: {tid}", 404)
            
            tank_num = tn_row._mapping.get("tank_number") if hasattr(tn_row, "_mapping") else tn_row[0]
            updates.append("tank_id = :tank_id")
            updates.append("tank_number = :tank_number")
            params["tank_id"] = tid
            params["tank_number"] = tank_num

        # --- Handle Standard Fields ---
        fields_to_update = [
            "inspection_date", "status_id", "inspection_type_id", "product_id", "location_id",
            "working_pressure", "frame_type", "design_temperature", "cabinet_type", "mfgr",
            "notes", "ownership", "safety_valve_brand_id", "safety_valve_model_id", "safety_valve_size_id",
            "vacuum_reading", "lifter_weight_value"
        ]

        for field in fields_to_update:
            if field in update_data:
                val = update_data[field]
                
                # Special Logic for Safety Valve Fields: Force invalid/empty to None
                if field in ["safety_valve_model_id", "safety_valve_size_id"]:
                    if not is_valid_id(val):
                        val = None
                
                # Special Logic for other IDs: Treat 0 as None if desired (based on user request)
                if field in ["status_id", "product_id", "inspection_type_id", "location_id", "safety_valve_brand_id"]:
                     if val == 0:
                         val = None

                updates.append(f"{field} = :{field}")
                params[field] = val

        # --- Execute Update & Validate ---
        if updates:
            sql = f"UPDATE tank_inspection_details SET {', '.join(updates)}, updated_at = NOW() WHERE inspection_id = :id"
            
            try:
                # Validation: Check if provided IDs exist (Only if they are NOT None)
                
                # Check Model
                if "safety_valve_model_id" in params:
                    mid = params["safety_valve_model_id"]
                    if mid is not None: # Strict None check
                        exists = db.execute(text("SELECT 1 FROM safety_valve_model WHERE id = :id LIMIT 1"), {"id": mid}).fetchone()
                        if not exists:
                            return error_resp(f"Invalid safety_valve_model_id: {mid}", 400)

                # Check Size
                if "safety_valve_size_id" in params:
                    sid = params["safety_valve_size_id"]
                    if sid is not None: # Strict None check
                        exists = db.execute(text("SELECT 1 FROM safety_valve_size WHERE id = :id LIMIT 1"), {"id": sid}).fetchone()
                        if not exists:
                            return error_resp(f"Invalid safety_valve_size_id: {sid}", 400)

                # Run Update
                db.execute(text(sql), params)
                db.commit()
                
            except Exception as e:
                db.rollback()
                logger.error(f"DB Error during update: {e}", exc_info=True)
                raise e

        return success_resp("Inspection details updated", {"inspection_id": inspection_id}, 200)

    except Exception as e:
        logger.error(f"Error updating tank inspection details {inspection_id}: {e}", exc_info=True)
        return error_resp("Error updating inspection details", 500)


# File: /mnt/data/tank_inspection_router.py
# Replace the existing @router.get("/review/{inspection_id}") handler with this complete function.


@router.get("/review/{inspection_id}")
def get_inspection_review(
    inspection_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        # ---------------------------------------------------------
        # PERMISSION CHECK
        # ---------------------------------------------------------
        row = db.execute(
            text("SELECT is_submitted FROM tank_inspection_details WHERE inspection_id = :id"),
            {"id": inspection_id}
        ).fetchone()

        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)

        is_submitted = int(row[0])
        role_id = current_user.role_id

        if is_submitted == 1 and role_id == 2:
            return error_resp("Access denied for submitted reports", 403)

        # ---------------------------------------------------------
        # 1. FETCH INSPECTION DETAILS
        # ---------------------------------------------------------
        inspection_row = db.execute(
            text("""
                SELECT 
                    ti.inspection_id,
                    ti.inspection_date,
                    ti.report_number,
                    ti.tank_id,
                    t.tank_number,
                    ti.status_id,
                    t.mfgr,
                    t.ownership,
                    t.cabinet_type,
                    ti.product_id,
                    ti.inspection_type_id,
                    ti.location_id,
                    ti.safety_valve_brand_id,
                    ti.safety_valve_model_id,
                    ti.safety_valve_size_id,
                    ti.working_pressure,
                    ti.design_temperature,
                    ti.frame_type,
                    pl.location_name,
                    pit.inspection_type_name,
                    pm.product_name,
                    ps.status_name,
                    sb.brand_name AS safety_valve_brand,
                    ti.vacuum_reading,
                    ti.lifter_weight_value,
                    ti.lifter_weight,
                    ti.is_reviewed,
                    ti.reviewed_by
                FROM tank_inspection_details ti
                LEFT JOIN tank_details t ON ti.tank_id = t.tank_id
                LEFT JOIN location_master pl ON ti.location_id = pl.id
                LEFT JOIN inspection_type pit ON ti.inspection_type_id = pit.id
                LEFT JOIN product_master pm ON ti.product_id = pm.id
                LEFT JOIN tank_status ps ON ti.status_id = ps.id
                LEFT JOIN safety_valve_brand sb ON ti.safety_valve_brand_id = sb.id
                WHERE ti.inspection_id = :iid
            """),
            {"iid": inspection_id}
        ).fetchone()

        if not inspection_row:
            return error_resp("Inspection not found", 404)

        inspection = dict(inspection_row._mapping)

        # ---------------------------------------------------------
        # 1.a FETCH next_inspection_date FROM tank_certificate
        # ---------------------------------------------------------
        try:
            cert_row = db.execute(
                text("""
                    SELECT next_insp_date
                    FROM tank_certificate
                    WHERE tank_number = :tank_number
                    ORDER BY id DESC
                    LIMIT 1
                """),
                {"tank_number": inspection.get("tank_number")}
            ).fetchone()

            inspection["next_inspection_date"] = (
                cert_row._mapping["next_insp_date"] if cert_row else None
            )

        except Exception as e:
            logger.warning(f"Failed to fetch next_inspection_date: {e}")
            inspection["next_inspection_date"] = None

        # ---------------------------------------------------------
        # LIFTER WEIGHT CDN URL
        # ---------------------------------------------------------
        from app.utils.s3_utils import to_cdn_url

        inspection["lifter_weight_url"] = (
            to_cdn_url(inspection["lifter_weight"])
            if inspection.get("lifter_weight")
            else None
        )

        # ---------------------------------------------------------
        # 2. FETCH IMAGES
        # ---------------------------------------------------------
        raw_images = db.execute(
            text("""
                SELECT image_type, image_path, thumbnail_path
                FROM tank_images
                WHERE inspection_id = :iid
            """),
            {"iid": inspection_id}
        ).fetchall()

        images = []
        for r in raw_images:
            img = dict(r._mapping)
            if str(img.get("image_type", "")).lower() == "lifter_weight":
                continue

            if img.get("image_path"):
                img["image_url"] = to_cdn_url(img["image_path"])
            if img.get("thumbnail_path"):
                img["thumbnail_url"] = to_cdn_url(img["thumbnail_path"])

            images.append(img)

        inspection["lifter_weight_thumbnail"] = None

        # ---------------------------------------------------------
        # 3. FETCH CHECKLIST
        # ---------------------------------------------------------
        checklist_rows = db.execute(
            text("""
                SELECT
                    ic.id,
                    ic.job_id,
                    ic.job_name,
                    ic.sub_job_id,
                    ic.sub_job_description,
                    ic.sn,
                    ic.status_id,
                    ic.comment,
                    s.status_name
                FROM inspection_checklist ic
                LEFT JOIN inspection_status s ON ic.status_id = s.status_id
                WHERE ic.inspection_id = :iid
                ORDER BY ic.job_id, ic.id
            """),
            {"iid": inspection_id}
        ).fetchall()

        from collections import OrderedDict
        job_groups = OrderedDict()

        for row in checklist_rows:
            r = dict(row._mapping)
            job_id = r["job_id"]

            if job_id not in job_groups:
                job_groups[job_id] = {
                    "job_id": job_id,
                    "title": r["job_name"],
                    "items": [],
                    "status_id": 1
                }

            if r["status_id"] == 2:
                job_groups[job_id]["status_id"] = 2
            elif r["status_id"] == 3:
                job_groups[job_id]["status_id"] = 1

            job_groups[job_id]["items"].append({
                "sn": r["sn"],
                "title": r["sub_job_description"],
                "sub_job_id": r["sub_job_id"],
                "status_id": r["status_id"],
                "status_name": r["status_name"],
                "comment": r["comment"]
            })

        inspection_checklist = list(job_groups.values())

        for job in inspection_checklist:
            sid = job["status_id"]
            job["status_name"] = (
                "OK" if sid == 1 else
                "NA" if sid == 2 else
                "Flagged" if sid == 3 else
                "Unknown"
            )

        # ---------------------------------------------------------
        # 4. FINAL RESPONSE
        # ---------------------------------------------------------
        return success_resp(
            "Inspection review fetched",
            {
                "inspection": inspection,
                "images": images,
                "inspection_checklist": inspection_checklist
            },
            200
        )

    except Exception as e:
        logger.exception("Inspection review error")
        return error_resp(str(e), 500)

@router.delete("/review/{inspection_id}")
def delete_inspection_review(inspection_id: int, db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    try:
        row = db.execute(text("SELECT * FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp("Inspection not found", 404)
        try:
            try:
                if hasattr(row, "_mapping"):
                    insp = dict(row._mapping)
                elif isinstance(row, dict):
                    insp = row
                else:
                    insp = dict((k, v) for k, v in row)
            except Exception:
                insp = jsonable_encoder(row)

            # Delete related checklist and to-do items (will cascade delete due to FK constraints)
            try:
                db.execute(text("DELETE FROM inspection_checklist WHERE inspection_id = :iid"), {"iid": str(inspection_id)})
                db.execute(text("DELETE FROM to_do_list WHERE inspection_id = :iid"), {"iid": str(inspection_id)})
            except Exception:
                db.rollback()
            db.execute(text("DELETE FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id})
            db.commit()
            return success_resp("Inspection and related checklist/to-do entries deleted", {"inspection_id": inspection_id}, 200)
        except Exception as e:
            db.rollback()
            logger.error(f"Error deleting review for {inspection_id}: {e}", exc_info=True)
            return error_resp(str(e), 500)
    except Exception as e:
        logger.error(f"Unexpected error deleting review for {inspection_id}: {e}", exc_info=True)
        return error_resp("Internal server error", 500)


# -------------------------
# Upload lifter weight (create/replace) endpoint
# -------------------------
@router.post("/{inspection_id}/lifter_weight", status_code=200)
def upload_lifter_weight(inspection_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    try:
        # Try to fetch thumbnail column as well
        try:
            row = db.execute(text("SELECT inspection_id, tank_number, lifter_weight, lifter_weight_thumbnail, is_submitted, web_submitted FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        except Exception:
            row = db.execute(text("SELECT inspection_id, tank_number, lifter_weight, is_submitted, web_submitted FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()

        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)
            
        # Submission check for role_id 2
        row_dict_check = dict(row._mapping) if hasattr(row, "_mapping") else dict(row)
        is_submitted = int(row_dict_check.get("is_submitted", 0))
        web_submitted = int(row_dict_check.get("web_submitted", 0))
        role_id = current_user.role_id
        
        if (is_submitted == 1 or web_submitted == 1) and role_id == 2:
            return error_resp("Cannot modify images for submitted inspection", 403)
        
        # Normalize row access
        thumb_rel = None
        if hasattr(row, "_mapping"):
            tank_number = row._mapping.get("tank_number")
            old_rel = row._mapping.get("lifter_weight")
            thumb_rel = row._mapping.get("lifter_weight_thumbnail") if "lifter_weight_thumbnail" in row._mapping else None
        else:
            try:
                old_rel = row[2]
                tank_number = row[1]
                if len(row) > 3:
                    thumb_rel = row[3]
            except Exception:
                old_rel = None
                tank_number = None

        if not file.content_type or not file.content_type.startswith("image/"):
            return error_resp("File must be an image", 400)

        saved = _save_lifter_file(file, tank_number, inspection_id)
        rel_path = saved["image_path"]
        thumb_path = saved.get("thumbnail_path")

        # Cleanup old files
        try:
            if old_rel:
                old_abs = os.path.join(IMAGES_ROOT_DIR, *old_rel.split("/"))
                if os.path.exists(old_abs):
                    try:
                        os.remove(old_abs)
                    except Exception:
                        logger.debug("Could not remove old lifter file: %s", old_abs, exc_info=True)
                
                # Cleanup old thumbnail (Explicit Path)
                if thumb_rel:
                    try:
                        thumb_abs = os.path.join(IMAGES_ROOT_DIR, *thumb_rel.split("/"))
                        if os.path.exists(thumb_abs):
                            os.remove(thumb_abs)
                    except Exception:
                        pass
                
                # Cleanup inferred thumbnail (if explicit path was missing but file exists in new structure)
                # old_rel might be "TANK/original/file.jpg" -> we check "TANK/thumbnail/file_thumb.jpg"
                try:
                    old_dir_name = os.path.dirname(old_abs) # .../original
                    old_file_name = os.path.basename(old_abs) # file.jpg
                    
                    # Check if we are in an 'original' folder
                    if os.path.basename(old_dir_name) == "originals":
                        base_dir = os.path.dirname(old_dir_name) # .../TANK
                        thumb_dir = os.path.join(base_dir, "thumbnails")
                        
                        # Construct expected thumbnail name
                        # Original: {tank}_{uuid}.jpg -> Thumb: {tank}_{uuid}_thumb.jpg
                        name_part, ext_part = os.path.splitext(old_file_name)
                        expected_thumb_name = f"{name_part}_thumb.jpg"
                        
                        expected_thumb_path = os.path.join(thumb_dir, expected_thumb_name)
                        if os.path.exists(expected_thumb_path):
                            try:
                                os.remove(expected_thumb_path)
                            except Exception:
                                pass
                except Exception:
                    pass

                # Legacy Cleanup (Same folder - for very old files)
                try:
                    old_base = os.path.splitext(os.path.basename(old_abs))[0]
                    folder = os.path.dirname(old_abs)
                    if os.path.isdir(folder):
                        for fn in os.listdir(folder):
                            if old_base in fn and "thumb" in fn:
                                try:
                                    os.remove(os.path.join(folder, fn))
                                except Exception:
                                    pass
                except Exception:
                    pass
        except Exception:
            logger.debug("Error while cleaning old lifter files for inspection %s", inspection_id, exc_info=True)

        # --- DB UPDATE (FIXED WITH THUMBNAIL) ---
        try:
            db.execute(text("""
                UPDATE tank_inspection_details 
                SET lifter_weight = :lp, 
                    lifter_weight_thumbnail = :thumb, 
                    updated_at = NOW() 
                WHERE inspection_id = :id
            """), {
                "lp": rel_path, 
                "thumb": thumb_path,
                "id": inspection_id
            })
            db.commit()
        except Exception as e_main:
            db.rollback()
            logger.warning(f"Failed to update with thumbnail column, trying without: {e_main}")
            # Fallback: maybe lifter_weight_thumbnail column doesn't exist?
            try:
                db.execute(text("""
                    UPDATE tank_inspection_details 
                    SET lifter_weight = :lp, 
                        updated_at = NOW() 
                    WHERE inspection_id = :id
                """), {
                    "lp": rel_path, 
                    "id": inspection_id
                })
                db.commit()
            except Exception as e_fallback:
                db.rollback()
                logger.error("Failed to update lifter_weight column (fallback also failed)", exc_info=True)
                return error_resp(f"Failed to save lifter weight path to DB: {e_fallback}", 500)

        return success_resp("Lifter weight photo uploaded", {"inspection_id": inspection_id, "lifter_weight": rel_path, "thumbnail": thumb_path}, 200)

    except Exception as e:
        logger.error(f"Error uploading lifter weight for inspection {inspection_id}: {e}", exc_info=True)
        return error_resp(str(e), 500)


@router.delete("/delete/inspection_details/{inspection_id}")
def delete_inspection_details(inspection_id: int, db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    try:
        row = db.execute(text("SELECT inspection_id FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)
        try:
            db.execute(text("DELETE FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id})
            db.commit()
            return success_resp("Inspection deleted", {"inspection_id": inspection_id}, 200)
        except Exception as e:
            db.rollback()
            logger.error(f"Error deleting inspection {inspection_id}: {e}", exc_info=True)
            return error_resp("Error deleting inspection", 500)
    except Exception as e:
        logger.error(f"Unexpected error deleting inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Internal server error", 500)


# -------------------------
# Tank details endpoint (keeps unfilled detection logic unchanged)
# -------------------------
@router.get("/tank-details/{tank_id}")
def get_tank_details(
    tank_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        # ---------------------------------------------------------
        # 1️⃣ Get tank_number from tank_details using tank_id
        # ---------------------------------------------------------
        tn_row = db.execute(
            text("""
                SELECT tank_number
                FROM tank_details
                WHERE tank_id = :tid
                LIMIT 1
            """),
            {"tid": tank_id}
        ).fetchone()

        if not tn_row:
            return error_resp(f"Tank not found for id: {tank_id}", 404)

        tank_number = (
            tn_row._mapping["tank_number"]
            if hasattr(tn_row, "_mapping")
            else tn_row[0]
        )

        if not tank_number:
            return error_resp("Tank number not found", 404)

        # ---------------------------------------------------------
        # 2️⃣ Fetch tank master details
        # ---------------------------------------------------------
        td_row = db.execute(
            text("""
                SELECT
                    working_pressure,
                    design_temperature,
                    frame_type,
                    cabinet_type,
                    mfgr,
                    ownership
                FROM tank_details
                WHERE tank_number = :tn
                LIMIT 1
            """),
            {"tn": tank_number}
        ).fetchone()

        if not td_row:
            return error_resp(f"Tank details not found: {tank_number}", 404)

        row = (
            td_row._mapping
            if hasattr(td_row, "_mapping")
            else {
                "working_pressure": td_row[0],
                "design_temperature": td_row[1],
                "frame_type": td_row[2],
                "cabinet_type": td_row[3],
                "mfgr": td_row[4],
                "ownership": td_row[5],
            }
        )

        # ---------------------------------------------------------
        # 3️⃣ Fetch PI NEXT INSPECTION DATE
        # ---------------------------------------------------------
        pi_row = db.execute(
            text("""
                SELECT next_insp_date
                FROM tank_certificate
                WHERE tank_number = :tn
                ORDER BY id DESC
                LIMIT 1
            """),
            {"tn": tank_number}
        ).fetchone()

        pi_next_inspection_date = None
        if pi_row:
            pi_next_inspection_date = (
                pi_row._mapping["next_insp_date"]
                if hasattr(pi_row, "_mapping")
                else pi_row[0]
            )

        # ---------------------------------------------------------
        # 4️⃣ Response (NO inspection_id here)
        # ---------------------------------------------------------
        def conv(v):
            return float(v) if isinstance(v, Decimal) else v

        data = {
            "tank_id": tank_id,
            "tank_number": tank_number,
            "working_pressure": conv(row.get("working_pressure")),
            "design_temperature": conv(row.get("design_temperature")),
            "frame_type": row.get("frame_type"),
            "cabinet_type": row.get("cabinet_type"),
            "mfgr": row.get("mfgr"),
            "ownership": row.get("ownership"),
            "pi_next_inspection_date": pi_next_inspection_date,
        }

        return success_resp("Tank details fetched", data, 200)

    except Exception as e:
        logger.error("Error fetching tank details", exc_info=True)
        return error_resp("Error fetching tank details", 500)
@router.get("/get/inspection/{inspection_id}")
def get_inspection_by_id(
    inspection_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user),
):
    """
    Fetch inspection record by inspection_id.

    RULES:
    - Inspection must belong to logged-in emp_id
    - is_submitted = 0  → return stored values
    - is_submitted = 1  → return empty values
    """

    try:
        # ---------------------------------------------------------
        # 1️⃣ Resolve emp_id from token
        # ---------------------------------------------------------
        emp_id_val = resolve_emp_id(current_user)
        if not emp_id_val:
            return error_resp("User authentication required", 401)

        role_id = current_user.role_id

        # ---------------------------------------------------------
        # 2️⃣ Fetch inspection row
        # ---------------------------------------------------------
        row = db.execute(
            text("""
                SELECT 
                    inspection_id,
                    tank_id,
                    tank_number,
                    report_number,
                    inspection_date,
                    status_id,
                    product_id,
                    inspection_type_id,
                    location_id,
                    working_pressure,
                    design_temperature,
                    frame_type,
                    cabinet_type,
                    mfgr,
                    pi_next_inspection_date,
                    safety_valve_brand_id,
                    safety_valve_model_id,
                    safety_valve_size_id,
                    notes,
                    created_by,
                    operator_id,
                    emp_id,
                    ownership,
                    lifter_weight,
                    vacuum_reading,
                    lifter_weight_value,
                    is_submitted,
                    web_submitted
                FROM tank_inspection_details
                WHERE inspection_id = :id
                LIMIT 1
            """),
            {"id": inspection_id},
        ).fetchone()

        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)

        # ---------------------------------------------------------
        # 3️⃣ Normalize row → dict
        # ---------------------------------------------------------
        if hasattr(row, "_mapping"):
            row_dict = dict(row._mapping)
        elif isinstance(row, dict):
            row_dict = row
        else:
            row_dict = dict(row)

        # ---------------------------------------------------------
        # 4️⃣ Ownership check (emp_id)
        # ---------------------------------------------------------
        try:
            row_emp_id = int(row_dict.get("emp_id"))
            req_emp_id = int(emp_id_val)
        except Exception:
            return error_resp("Invalid emp_id mapping", 403)

        if row_emp_id != req_emp_id:
            logger.warning(
                f"Inspection {inspection_id} access denied: "
                f"belongs to emp_id={row_emp_id}, requested by emp_id={req_emp_id}"
            )
            return error_resp("Inspection not found or access denied", 404)

        # ---------------------------------------------------------
        # 5️⃣ Submission logic with role-based access
        # ---------------------------------------------------------
        is_submitted = int(row_dict.get("is_submitted", 0))
        web_submitted = int(row_dict.get("web_submitted", 0))

        if is_submitted == 1 or web_submitted == 1:
            if role_id == 2:
                return error_resp("Access denied for submitted reports", 403)
            elif role_id not in [1, 3, 4]:
                # Return empty values for editing
                return success_resp(
                    "Inspection already submitted",
                    {
                        "inspection_id": None,
                        "tank_id": row_dict.get("tank_id"),
                        "status_id": None,
                        "product_id": None,
                        "inspection_type_id": None,
                        "location_id": None,
                        "safety_valve_brand_id": None,
                        "safety_valve_model_id": None,
                        "safety_valve_size_id": None,
                        "notes": None,
                        "operator_id": None,
                        "inspection_date": None,
                        "report_number": None,
                        "lifter_weight": None,
                        "vacuum_reading": None,
                        "lifter_weight_value": None,
                        "is_submitted": 1,
                    },
                    200,
                )
            # For role 1,3,4, fall through to return full values

        # 🟢 NOT SUBMITTED or ALLOWED SUBMITTED → return ACTUAL STORED VALUES
        # Fetch fresh tank details (to display authoritative values in report header)
        try:
            td = fetch_tank_details(db, row_dict.get("tank_number"))
        except Exception:
            td = {}

        # Also fetch latest PI next inspection date from tank_certificate (if available)
        try:
            pi_row2 = db.execute(
                text(
                    """
                    SELECT next_insp_date
                    FROM tank_certificate
                    WHERE tank_number = :tn
                    ORDER BY id DESC
                    LIMIT 1
                    """
                ),
                {"tn": row_dict.get("tank_number")},
            ).fetchone()
            if pi_row2:
                if hasattr(pi_row2, "_mapping"):
                    td_pi = pi_row2._mapping.get("next_insp_date")
                elif isinstance(pi_row2, dict):
                    td_pi = pi_row2.get("next_insp_date")
                else:
                    td_pi = pi_row2[0]
            else:
                td_pi = None
        except Exception:
            td_pi = None

        def conv(v):
            return float(v) if isinstance(v, Decimal) else v

        resp = {
            "inspection_id": row_dict.get("inspection_id"),
            "tank_id": row_dict.get("tank_id"),
            "status_id": row_dict.get("status_id"),
            "product_id": row_dict.get("product_id"),
            "inspection_type_id": row_dict.get("inspection_type_id"),
            "location_id": row_dict.get("location_id"),
            "safety_valve_brand_id": row_dict.get("safety_valve_brand_id"),
            "safety_valve_model_id": row_dict.get("safety_valve_model_id"),
            "safety_valve_size_id": row_dict.get("safety_valve_size_id"),
            "notes": row_dict.get("notes"),
            "operator_id": row_dict.get("operator_id"),
            "inspection_date": row_dict.get("inspection_date"),
            "report_number": row_dict.get("report_number"),
            "lifter_weight": row_dict.get("lifter_weight"),
            "vacuum_reading": row_dict.get("vacuum_reading"),
            "lifter_weight_value": row_dict.get("lifter_weight_value"),
            "is_submitted": is_submitted,
            # -- authoritative tank details (sourced from tank_details table)
            "working_pressure": conv(td.get("working_pressure") if td else row_dict.get("working_pressure")),
            "design_temperature": td.get("design_temperature") if td else row_dict.get("design_temperature"),
            "frame_type": td.get("frame_type") if td else row_dict.get("frame_type"),
            "cabinet_type": td.get("cabinet_type") if td else row_dict.get("cabinet_type"),
            "mfgr": td.get("mfgr") if td else row_dict.get("mfgr"),
            "ownership": td.get("ownership") if td else row_dict.get("ownership"),
            "pi_next_inspection_date": td_pi if td_pi is not None else row_dict.get("pi_next_inspection_date"),
        }

        return success_resp("Inspection fetched successfully", resp, 200)

    except Exception as e:
        logger.error(f"Error fetching inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Internal server error", 500)

# -------------------------
@router.get("/inspection/latest-draft/{tank_id}")
def get_latest_draft_inspection(
    tank_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user),
):
    """
    Return the latest unsubmitted inspection_id for the given tank_id
    belonging to the current emp_id, or inspection_id = None if no draft exists.
    """
    try:
        emp_id_val = resolve_emp_id(current_user)
        if not emp_id_val:
            return error_resp("User authentication required", 401)

        row = db.execute(
            text(
                """
                SELECT inspection_id
                FROM tank_inspection_details
                WHERE tank_id = :tid
                  AND emp_id = :eid
                  AND is_submitted = 0
                ORDER BY inspection_id DESC
                LIMIT 1
                """
            ),
            {"tid": tank_id, "eid": emp_id_val},
        ).fetchone()

        inspection_id = None
        if row:
            if hasattr(row, "_mapping"):
                inspection_id = row._mapping.get("inspection_id")
            elif isinstance(row, dict):
                inspection_id = row.get("inspection_id")
            else:
                inspection_id = row[0]

        return success_resp(
            "Latest draft inspection fetched",
            {"inspection_id": inspection_id},
            200,
        )
    except Exception as e:
        logger.error(f"Error fetching latest draft inspection for tank {tank_id}: {e}", exc_info=True)
        return error_resp("Error fetching latest draft inspection", 500)
# -------------------------
# Delete lifter weight endpoint (keeps same semantics)
# -------------------------
@router.delete("/{inspection_id}/lifter_weight", status_code=200)
def delete_lifter_weight(inspection_id: int, db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    try:
        # Try to fetch thumbnail column as well
        try:
            row = db.execute(text("SELECT inspection_id, tank_number, lifter_weight, lifter_weight_thumbnail, is_submitted, web_submitted FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        except Exception:
            row = db.execute(text("SELECT inspection_id, tank_number, lifter_weight, is_submitted, web_submitted FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()

        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)
            
        # Submission check for role_id 2
        row_dict_check = dict(row._mapping) if hasattr(row, "_mapping") else dict(row)
        is_submitted = int(row_dict_check.get("is_submitted", 0))
        web_submitted = int(row_dict_check.get("web_submitted", 0))
        role_id = current_user.role_id
        
        if (is_submitted == 1 or web_submitted == 1) and role_id == 2:
            return error_resp("Cannot delete images for submitted inspection", 403)

        thumb_rel = None
        if hasattr(row, "_mapping"):
            rel = row._mapping.get("lifter_weight")
            thumb_rel = row._mapping.get("lifter_weight_thumbnail") if "lifter_weight_thumbnail" in row._mapping else None
        else:
            try:
                rel = row[2]
                if len(row) > 3:
                    thumb_rel = row[3]
            except Exception:
                rel = None

        if not rel:
            return error_resp("No lifter weight image present for this inspection", 404)

        # Delete Original
        try:
            abs_path = os.path.join(IMAGES_ROOT_DIR, *rel.split("/"))
            if os.path.exists(abs_path):
                try:
                    os.remove(abs_path)
                except Exception:
                    logger.debug("Could not remove lifter file %s", abs_path, exc_info=True)
        except Exception:
             pass

        # Delete Thumbnail (Explicit Path)
        if thumb_rel:
            try:
                thumb_abs = os.path.join(IMAGES_ROOT_DIR, *thumb_rel.split("/"))
                if os.path.exists(thumb_abs):
                    os.remove(thumb_abs)
            except Exception:
                pass
        
        # Legacy Thumbnail Cleanup (Same folder)
        try:
            folder = os.path.dirname(abs_path)
            base_no_ext = os.path.splitext(os.path.basename(abs_path))[0]
            if os.path.isdir(folder):
                for fn in os.listdir(folder):
                    if base_no_ext in fn and "thumb" in fn:
                        try:
                            os.remove(os.path.join(folder, fn))
                        except Exception:
                            pass
        except Exception:
            pass

        try:
            # Try to nullify both columns
            try:
                db.execute(text("UPDATE tank_inspection_details SET lifter_weight = NULL, lifter_weight_thumbnail = NULL, updated_at = NOW() WHERE inspection_id = :id"), {"id": inspection_id})
            except Exception:
                db.execute(text("UPDATE tank_inspection_details SET lifter_weight = NULL, updated_at = NOW() WHERE inspection_id = :id"), {"id": inspection_id})
            
            db.commit()
        except Exception:
            db.rollback()
            logger.error("Failed to clear lifter_weight DB column", exc_info=True)
            return error_resp("Failed to remove lifter weight reference from DB", 500)

        return success_resp("Lifter weight image deleted", {"inspection_id": inspection_id}, 200)
    except Exception as e:
        logger.error(f"Error deleting lifter weight for inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Error deleting lifter weight", 500)
    
@router.put("/{inspection_id}/lifter_weight")
def update_lifter_weight(inspection_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    """
    Update (Replace) the lifter weight image for an inspection.
    This logic calls the existing upload function but exposes it via PUT
    for semantic correctness (Replace existing resource).
    """
    # We reuse the logic from the existing POST function
    return upload_lifter_weight(inspection_id, file, db)

# ----------------------------
# SUBMIT INSPECTION (Finalize)
# ----------------------------
@router.get("/submit")
def submit_inspection(
    inspection_id: int = Header(..., alias="inspection-id"),
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Finalize the inspection.
    First validates that all required data is complete (no null values, all images present, to_do_list empty).
    Only submits if validation passes.
    Sets the status_id to 4 (Completed), is_submitted to 1, and updates the timestamp.
    """
# THEN run validation
# THEN update status_id / is_submitted

    try:
        # 1. Verify Inspection Exists
        row = db.execute(text("SELECT inspection_id FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)

        # 2. Run Validation Check (same logic as validation endpoint)
        issues = {"inspection": [], "checklist": [], "to_do_list": [], "images": []}
        
        # 2a. Check inspection row
        try:
            insp_row = db.execute(text("SELECT * FROM tank_inspection_details WHERE inspection_id = :id LIMIT 1"), {"id": inspection_id}).fetchone()
            if hasattr(insp_row, "_mapping"):
                insp = dict(insp_row._mapping)
            elif isinstance(insp_row, dict):
                insp = insp_row
            else:
                try:
                    insp = dict(zip(insp_row.keys(), insp_row))
                except Exception:
                    insp = {}

            required_inspection_fields = [
                "tank_id", "tank_number", "report_number", "inspection_date",
                "status_id", "product_id", "inspection_type_id", "location_id",
                "vacuum_reading", "lifter_weight_value",
            ]

            for f in required_inspection_fields:
                v = insp.get(f)
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    issues["inspection"].append({"field": f, "reason": "null or empty"})
                else:
                    if isinstance(v, (int, float)) and int(v) == 0:
                        issues["inspection"].append({"field": f, "reason": "zero or invalid"})

            # Validate PI next inspection date
            pi_keys = ["pi_next_inspection_date", "pi_next_insp_date", "next_insp_date", "pi_nextinsp_date"]
            pi_found = False
            for k in pi_keys:
                v = insp.get(k)
                if v is not None and not (isinstance(v, str) and v.strip() == ""):
                    pi_found = True
                    break
            # If not found in inspection, check tank_certificate
            if not pi_found:
                tank_id = insp.get("tank_id")
                if tank_id:
                    try:
                        row = db.execute(
                            text(
                                """
                                SELECT next_insp_date
                                FROM tank_certificate
                                WHERE tank_id = :tank_id
                                ORDER BY next_insp_date IS NULL ASC, next_insp_date DESC
                                LIMIT 1
                                """
                            ),
                            {"tank_id": tank_id},
                        ).fetchone()
                        if row and row[0]:
                            pi_found = True
                    except Exception:
                        pass
            if not pi_found:
                issues["inspection"].append({"field": "pi_next_inspection_date", "reason": "null or empty"})

        except Exception as e:
            logger.exception("Error validating inspection: %s", e)
            return error_resp(f"Error validating inspection: {e}", 500)

        # 2b. Validate inspection_checklist
        try:
            checklist_rows = db.execute(text("SELECT * FROM inspection_checklist WHERE inspection_id = :id"), {"id": str(inspection_id)}).fetchall() or []
            if not checklist_rows:
                issues["checklist"].append({"reason": "no checklist rows found for this inspection"})
            else:
                for r in checklist_rows:
                    rr = dict(r._mapping) if hasattr(r, "_mapping") else dict(zip(r.keys(), r))
                    row_issue = {"id": rr.get("id")}
                    for f in ("job_id", "sub_job_id", "sn", "status_id"):
                        v = rr.get(f)
                        if v is None or (isinstance(v, str) and v.strip() == ""):
                            row_issue.setdefault("missing_fields", []).append(f)
                    if "missing_fields" in row_issue:
                        issues["checklist"].append(row_issue)
        except Exception as e:
            logger.exception("Error validating checklist: %s", e)
            return error_resp(f"Error validating checklist: {e}", 500)

        # 2c. Validate to_do_list is empty
        try:
            todo_rows = db.execute(text("""
                SELECT DISTINCT c.job_id, c.job_name, t.status_id
                FROM to_do_list t
                LEFT JOIN inspection_checklist c ON t.checklist_id = c.id
                WHERE t.inspection_id = :id AND t.status_id = 2
                ORDER BY c.job_id
            """), {"id": inspection_id}).fetchall() or []
            
            if todo_rows:
                flagged_jobs = []
                for r in todo_rows:
                    rr = dict(r._mapping) if hasattr(r, "_mapping") else dict(zip(r.keys(), r))
                    job_id = rr.get("job_id")
                    job_name = rr.get("job_name")
                    if job_id is not None:
                        flagged_jobs.append({
                            "job_id": str(job_id),
                            "job_name": job_name or "",
                            "status_id": 2
                        })
                
                if flagged_jobs:
                    issues["to_do_list"] = [{
                        "reason": "to_do_list not empty - inspection has flagged items",
                        "flagged_jobs": flagged_jobs
                    }]
        except Exception as e:
            logger.exception("Error validating to_do_list: %s", e)

        # 2d. Validate images
        try:
            img_rows = db.execute(text("SELECT image_type, image_path, thumbnail_path, image_id FROM tank_images WHERE inspection_id = :id"), {"id": inspection_id}).fetchall() or []
            img_count = len(img_rows)
            
            expected_types = db.execute(text("SELECT id, image_type, count FROM image_type")).fetchall() or []
            expected_total_images = 0
            for et in expected_types:
                if hasattr(et, "_mapping"):
                    cnt = et._mapping.get("count") or 1
                elif isinstance(et, dict):
                    cnt = et.get("count") or 1
                else:
                    try:
                        _, _, cnt = et
                    except Exception:
                        cnt = 1
                expected_total_images += int(cnt)

            if expected_total_images == 0:
                expected_total_images = 15
            
            if img_count < expected_total_images:
                issues["images"].append({"reason": f"insufficient images: found {img_count}, expected {expected_total_images}"})
            else:
                for idx, r in enumerate(img_rows):
                    rr = dict(r._mapping) if hasattr(r, "_mapping") else dict(zip(r.keys(), r))
                    if not rr.get("image_path"):
                        issues["images"].append({"index": idx, "reason": "image_path missing"})
                    if (not rr.get("image_id")) and (not rr.get("image_type")):
                        issues["images"].append({"index": idx, "reason": "image type missing"})
        except Exception as e:
            logger.exception("Error validating images: %s", e)

        # 3. Check if any issues found
        any_issues = any(len(v) > 0 for v in issues.values())
        if any_issues:
            # Return detailed validation errors
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "message": "Cannot submit inspection - validation failed. Please complete all required fields.",
                    "data": {
                        "inspection_id": inspection_id,
                        "issues": issues
                    }
                }
            )

        # 4. Update Status to 4 (Completed) and is_submitted = 1
        # 4. Update Status logic based on role_id
        role_id = current_user.role_id
        
        if role_id in [1, 3, 4]:
            # For Admin/Mgmt: web_submitted = 1, is_submitted remains 0
            # status_id is NOT finalized to 4 yet (because is_submitted is not 1)
             db.execute(text("""
                UPDATE tank_inspection_details 
                SET web_submitted = 1, updated_at = NOW() 
                WHERE inspection_id = :id
            """), {"id": inspection_id})
             msg = "Inspection web-submitted successfully"

        elif role_id == 2:
            # For Operator: is_submitted = 1, web_submitted = 0
            # update status to Completed (4) as well if that was the original logic for operators?
            # User said: "is_sumitted tuns to 1 AND WEB_SUBMITTED = 0"
            # And previously we had status_id = 4. Let's keep status_id updates out if user didn't ask, 
            # BUT the previous code set status_id=4. The user said specifically what to set.
            # I will follow strictly: "is_sumitted tuns to 1 AND WEB_SUBMITTED = 0"
            # The previous status_id=4 change was removed in step 561 per user request.
            
            db.execute(text("""
                UPDATE tank_inspection_details 
                SET is_submitted = 1, web_submitted = 0, updated_at = NOW() 
                WHERE inspection_id = :id
            """), {"id": inspection_id})
            msg = "Inspection submitted successfully"
            
        else:
             # Fallback for unknown roles (treat as operator or error? Defaulting to is_submitted=1 for safety if that was old behavior)
             db.execute(text("""
                UPDATE tank_inspection_details 
                SET is_submitted = 1, updated_at = NOW() 
                WHERE inspection_id = :id
            """), {"id": inspection_id})
             msg = "Inspection submitted successfully"
        
        db.commit()
        
        return success_resp(msg, {"inspection_id": inspection_id, "status": "Submitted"}, 200)

    except Exception as e:
        db.rollback()
        logger.error(f"Error submitting inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Failed to submit inspection", 500)

# ----------------------------
# REVIEW INSPECTION (Mark reviewed)
# ----------------------------
@router.post("/review_finalize/{inspection_id}")
def review_finalize_inspection(
    inspection_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Mark the inspection as reviewed.
    Sets is_reviewed = 1 and reviewed_by = current_user.emp_id.
    """
    emp_id_val = resolve_emp_id(current_user)
    if not emp_id_val:
        return error_resp("Unauthorized", 401)

    try:
        # Check if inspection exists
        row = db.execute(text("SELECT inspection_id FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)

        # Update is_reviewed and reviewed_by
        db.execute(
            text("UPDATE tank_inspection_details SET is_reviewed = 1, reviewed_by = :reviewer, updated_at = NOW() WHERE inspection_id = :id"),
            {"reviewer": emp_id_val, "id": inspection_id}
        )
        db.commit()

        # Insert into inspection_history
        try:
            # Fetch the updated record
            record = db.execute(text("SELECT * FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
            if record:
                history_entry = InspectionHistory(
                    inspection_id=record.inspection_id,
                    inspection_date=record.inspection_date,
                    created_at=record.created_at,
                    updated_at=record.updated_at,
                    report_number=record.report_number,
                    tank_id=record.tank_id,
                    tank_number=record.tank_number,
                    status_id=record.status_id,
                    product_id=record.product_id,
                    inspection_type_id=record.inspection_type_id,
                    location_id=record.location_id,
                    working_pressure=record.working_pressure,
                    design_temperature=record.design_temperature,
                    frame_type=record.frame_type,
                    cabinet_type=record.cabinet_type,
                    mfgr=record.mfgr,
                    safety_valve_brand_id=record.safety_valve_brand_id,
                    safety_valve_model_id=record.safety_valve_model_id,
                    safety_valve_size_id=record.safety_valve_size_id,
                    pi_next_inspection_date=record.pi_next_inspection_date,
                    notes=record.notes,
                    lifter_weight=record.lifter_weight,
                    lifter_weight_thumbnail=record.lifter_weight_thumbnail,
                    vacuum_reading=record.vacuum_reading,
                    lifter_weight_value=record.lifter_weight_value,
                    emp_id=record.emp_id,
                    operator_id=record.operator_id,
                    ownership=record.ownership,
                    is_submitted=record.is_submitted,
                    is_reviewed=record.is_reviewed,
                    reviewed_by=record.reviewed_by,
                    created_by=record.created_by,
                    updated_by=record.updated_by,
                    history_date=func.now()
                )
                db.add(history_entry)
                db.commit()
        except Exception as e:
            logger.error(f"Error inserting into inspection_history for {inspection_id}: {e}")
            # Don't fail the whole operation

        return success_resp("Inspection report marked as REVIEWED", {"inspection_id": inspection_id, "reviewed_by": emp_id_val}, 200)

    except Exception as e:
        db.rollback()
        logger.error(f"Error reviewing inspection {inspection_id}: {e}", exc_info=True)
        return error_resp(str(e), 500)

# ----------------------------
# GET INSPECTION HISTORY
# ----------------------------
@router.get("/history")
def get_inspection_history(
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Get all inspection history records.
    """
    try:
        role_id = current_user.role_id
        if role_id == 2:
            return success_resp("History", [], 200)

        results = db.execute(text("""
            SELECT 
                ih.id,
                ih.inspection_id,
                ih.inspection_date,
                ih.report_number,
                ih.tank_number,
                ih.history_date,
                ih.is_reviewed,
                ih.reviewed_by,
                ps.status_name,
                pit.inspection_type_name,
                pm.product_name,
                pl.location_name
            FROM inspection_history ih
            LEFT JOIN tank_status ps ON ih.status_id = ps.id
            LEFT JOIN inspection_type pit ON ih.inspection_type_id = pit.id
            LEFT JOIN product_master pm ON ih.product_id = pm.id
            LEFT JOIN location_master pl ON ih.location_id = pl.id
            ORDER BY ih.history_date DESC
        """)).fetchall()

        history = [dict(r._mapping) for r in results]

        return success_resp("Inspection history fetched", history, 200)

    except Exception as e:
        logger.exception("Error fetching inspection history")
        return error_resp("Failed to fetch history", 500)

# -----------------------------
# GET CURRENT USER INFO
# -----------------------------
@router.get("/user/me")
def get_current_user_info(current_user: Optional[dict] = Depends(get_current_user)):
    """Return basic info about the currently authenticated user.

    Accepts either a `User` model instance or a decoded payload dict.
    """
    if not current_user:
        return error_resp("User not authenticated", 401)

    if hasattr(current_user, "emp_id"):
        emp_id = getattr(current_user, "emp_id", None)
        role_id = getattr(current_user, "role_id", None)
        login_name = getattr(current_user, "login_name", None)
        email = getattr(current_user, "email", None)
    else:
        emp_id = current_user.get("emp_id")
        role_id = current_user.get("role_id")
        login_name = current_user.get("login_name")
        email = current_user.get("email")

    return success_resp("User info", {
        "emp_id": emp_id,
        "role_id": role_id,
        "login_name": login_name,
        "email": email,
    })

@router.post("/copy/{inspection_id}")
def copy_inspection(
    inspection_id: int,
    new_type_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        from app.services.copy_inspection_service import CopyInspectionService
        
        new_id, new_report_no = CopyInspectionService.copy_inspection(
            db, inspection_id, new_type_id, current_user
        )
        
        return success_resp("Inspection copied successfully", {
            "new_inspection_id": new_id,
            "new_report_number": new_report_no
        }, 201)

    except ValueError as e:
        return error_resp(str(e), 404)
    except Exception as e:
        logger.error(f"Error copying inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Failed to copy inspection", 500)

